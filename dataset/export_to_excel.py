"""Convert the sample SQuAD dataset to a readable Excel workbook.

Produces dataset/sample_squad.xlsx with two sheets:
  1. "QA Examples" - one row per question with context, question, answer,
     answer_start character offset, passage title.
  2. "Passages"    - the 8 unique passages with their titles.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "dataset" / "sample_squad.json"
OUT = ROOT / "dataset" / "sample_squad.xlsx"

# ---- load source ----------------------------------------------------------
with open(SRC, "r", encoding="utf-8") as f:
    squad = json.load(f)

# ---- styling --------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

# ---- Sheet 1: QA Examples ---------------------------------------------------
ws = wb.active
ws.title = "QA Examples"
cols = ["Passage Title", "Context", "Question", "Answer", "Answer Start (char)", "Answer End (char)"]
ws.append(cols)
for c, _ in enumerate(cols, 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center")

row = 2
for article in squad["data"]:
    title = article["title"]
    for para in article["paragraphs"]:
        context = para["context"]
        for qa in para["qas"]:
            ans = qa["answers"][0]
            start = ans["answer_start"]
            ws.append([title, context, qa["question"], ans["text"], start, start + len(ans["text"])])
            for c in range(1, 7):
                ws.cell(row=row, column=c).alignment = WRAP
            row += 1

# ---- Sheet 2: Passages -------------------------------------------------------
ws2 = wb.create_sheet("Passages")
cols2 = ["Passage Title", "Context", "Word Count", "Char Count"]
ws2.append(cols2)
for c, _ in enumerate(cols2, 1):
    cell = ws2.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center")

row = 2
for article in squad["data"]:
    for para in article["paragraphs"]:
        context = para["context"]
        words = len(context.split())
        ws2.append([article["title"], context, words, len(context)])
        ws2.cell(row=row, column=2).alignment = WRAP
        row += 1

# ---- column widths ------------------------------------------------------------
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 70
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 100
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12

wb.save(OUT)
print(f"[ok] wrote {OUT}")
print(f"     QA rows : {ws.max_row - 1}")
print(f"     Passage rows: {ws2.max_row - 1}")
